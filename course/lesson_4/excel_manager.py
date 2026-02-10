from typing import Any

import openpyxl
from openpyxl.workbook import Workbook


class ReadExcelFile:
    file_path: str = "store_simple.xlsx"

    @staticmethod
    def read_all_from_excel(
        file_path: str = file_path,
    ) -> list[Any] | tuple[list, list]:

        wb: Workbook = openpyxl.load_workbook(file_path)
        sheet = wb.active
        quantity_in_stock_list: list = []
        row: int = 2
        while sheet.cell(row=row, column=1).value:
            quantity_in_stock_list.append(
                (
                    sheet.cell(row=row, column=1).value,
                    sheet.cell(row=row, column=2).value,
                    sheet.cell(row=row, column=3).value,
                    sheet.cell(row=row, column=4).value,
                )
            )
            row += 1

        customer_start = None
        for r in range(1, sheet.max_row + 1):
            if sheet.cell(row=r, column=1).value == "Customer Order":
                customer_start = r + 2
                break

        if not customer_start:
            return []

        customer_order_list: list = []
        row: int = customer_start
        while sheet.cell(row=row, column=1).value:
            customer_order_list.append(
                (
                    [
                        sheet.cell(row=row, column=1).value,
                        sheet.cell(row=row, column=2).value,
                    ],
                    [
                        sheet.cell(row=row, column=3).value,
                        sheet.cell(row=row, column=4).value,
                        sheet.cell(row=row, column=5).value,
                        sheet.cell(row=row, column=6).value,
                    ],
                )
            )
            row += 1

        return quantity_in_stock_list, customer_order_list
